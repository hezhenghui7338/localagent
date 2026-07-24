"""Document loaders for supported kb/ file types."""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path

from localagent import config
from localagent.config import IMAGE_SUFFIXES, SUPPORTED_SUFFIXES


@dataclass
class LoadedDoc:
    text: str
    source: str
    filename: str
    metadata: dict = field(default_factory=dict)


def _load_txt(path: Path) -> str:
    return path.read_text(encoding="utf-8", errors="replace")


def _load_xlsx(path: Path) -> str:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    parts: list[str] = []
    for sheet in wb.worksheets:
        parts.append(f"## Sheet: {sheet.title}")
        for row in sheet.iter_rows(values_only=True):
            cells = [str(c).strip() for c in row if c is not None and str(c).strip()]
            if cells:
                parts.append(" | ".join(cells))
    wb.close()
    return "\n".join(parts)


def _pdf_needs_ocr(page_count: int, pages_with_text: int) -> bool:
    if page_count <= 0:
        return False
    if pages_with_text <= 0:
        return True
    ratio = pages_with_text / page_count
    return ratio < config.OCR_PDF_TEXT_RATIO


def _load_pdf(path: Path) -> tuple[str, dict]:
    """Extract text per page; inject ## [p.N] markers for citeable summarize."""
    from pypdf import PdfReader

    reader = PdfReader(str(path))
    parts: list[str] = []
    page_count = len(reader.pages)
    non_empty = 0
    for index, page in enumerate(reader.pages, start=1):
        try:
            raw = page.extract_text() or ""
        except Exception:
            raw = ""
        text = raw.strip()
        if not text:
            continue
        non_empty += 1
        parts.append(f"## [p.{index}]\n{text}")
    meta = {"page_count": page_count, "pages_with_text": non_empty}
    body = "\n\n".join(parts)

    if body.strip() or not _pdf_needs_ocr(page_count, non_empty):
        return body, meta

    if not config.OCR_ENABLED:
        meta["needs_ocr"] = True
        return "", meta

    from localagent.ingest.ocr import ocr_metadata_from_result, ocr_pdf

    ocr_result = ocr_pdf(path)
    meta.update(ocr_metadata_from_result(ocr_result))
    return ocr_result.text, meta


def _load_image(path: Path) -> tuple[str, dict]:
    if config.OCR_ENABLED:
        from localagent.ingest.ocr import ocr_image, ocr_metadata_from_result

        ocr_result = ocr_image(path)
        return ocr_result.text, ocr_metadata_from_result(ocr_result)

    if config.VL_ENABLED:
        from localagent.ingest.vision import caption_image

        return caption_image(path), {"vl_caption": True}

    return "", {"needs_ocr": True}


def explain_load_failure(path: Path) -> str:
    """Return a user-facing hint when ``load_file`` would return None."""
    from localagent.ingest.ocr import ocr_install_hint

    path = Path(path)
    suffix = path.suffix.lower()

    if suffix == ".pdf":
        try:
            _, meta = _load_pdf_text_layer_only(path)
        except Exception:
            return f"无法读取 PDF: {path}"
        page_count = int(meta.get("page_count") or 0)
        pages_with_text = int(meta.get("pages_with_text") or 0)
        if _pdf_needs_ocr(page_count, pages_with_text):
            if config.OCR_ENABLED:
                return f"扫描版 PDF OCR 失败或内容为空: {path}"
            return f"扫描版 PDF 无文本层。{ocr_install_hint()}"
        return f"无法读取 PDF 内容: {path}"

    if suffix in IMAGE_SUFFIXES:
        if not config.OCR_ENABLED and not config.VL_ENABLED:
            return f"图片需要 OCR 或 VL 才能读取。{ocr_install_hint()}"
        if config.OCR_ENABLED:
            return f"图片 OCR 失败或内容为空: {path}"
        return f"图片 VL 描述失败或内容为空: {path}"

    return f"无法读取文件内容: {path}"


def _load_pdf_text_layer_only(path: Path) -> tuple[str, dict]:
    """Inspect PDF text layer without triggering OCR."""
    from pypdf import PdfReader

    reader = PdfReader(str(path))
    parts: list[str] = []
    page_count = len(reader.pages)
    non_empty = 0
    for index, page in enumerate(reader.pages, start=1):
        try:
            raw = page.extract_text() or ""
        except Exception:
            raw = ""
        text = raw.strip()
        if not text:
            continue
        non_empty += 1
        parts.append(f"## [p.{index}]\n{text}")
    meta = {"page_count": page_count, "pages_with_text": non_empty}
    return "\n\n".join(parts), meta


def load_file(path: Path) -> LoadedDoc | None:
    path = Path(path)
    if not path.exists() or not path.is_file():
        return None

    suffix = path.suffix.lower()
    if suffix not in SUPPORTED_SUFFIXES:
        return None

    extra_meta: dict = {}
    if suffix in {".md", ".markdown", ".txt"}:
        text = _load_txt(path)
    elif suffix == ".xlsx":
        text = _load_xlsx(path)
    elif suffix == ".pdf":
        text, extra_meta = _load_pdf(path)
    elif suffix in IMAGE_SUFFIXES:
        text, extra_meta = _load_image(path)
    else:
        return None

    text = text.strip()
    if not text:
        return None

    modified_at = datetime.fromtimestamp(path.stat().st_mtime).isoformat(timespec="seconds")
    metadata = {"suffix": suffix, "modified_at": modified_at, **extra_meta}

    return LoadedDoc(
        text=text,
        source=str(path.resolve()),
        filename=path.name,
        metadata=metadata,
    )
