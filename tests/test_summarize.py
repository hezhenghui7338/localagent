"""Tests for one-click document summarize (short path)."""

from __future__ import annotations

from pathlib import Path

import pytest

from localagent import config
from localagent.i18n import reset_lang_cache
from localagent.summarize.document import (
    KEEP_HINT,
    DocumentTooLongError,
    citation_ok,
    ensure_citations,
    not_kept_hint_if_asked,
    summarize_loaded,
    summarize_path,
)
from localagent.ingest.loader import LoadedDoc, load_file


@pytest.fixture(autouse=True)
def _force_zh_ui_lang(monkeypatch):
    """Summarize UI strings stay Chinese in assertions; pin LA_LANG for CI locales."""
    monkeypatch.setenv("LA_LANG", "zh")
    reset_lang_cache()
    yield
    reset_lang_cache()


def _md_doc(tmp_path: Path) -> Path:
    path = tmp_path / "brief.md"
    path.write_text(
        "# 产品概述\n\n"
        "LocalAgent 是跑在本机上的个人 AI 助手。\n\n"
        "## 安装\n\n"
        "使用 pipx 安装后运行 la setup。\n\n"
        "## 限制\n\n"
        "本版本不支持扫描版 PDF 的 OCR。\n",
        encoding="utf-8",
    )
    return path


def test_heuristic_summarize_md_has_citations(tmp_path: Path):
    path = _md_doc(tmp_path)
    result = summarize_path(path, keep=False, use_llm=False)
    assert "## 总结" in result.markdown
    assert "## 结构化要点" in result.markdown
    assert "§" in result.markdown or "〔" in result.markdown
    assert result.kept is False
    assert result.used_llm is False


def test_summarize_default_not_kept(tmp_path: Path):
    path = _md_doc(tmp_path)
    before = list(config.KB_DIR.iterdir()) if config.KB_DIR.exists() else []
    result = summarize_path(path, keep=False, use_llm=False)
    assert result.kept is False
    assert result.keep_target is None
    after = list(config.KB_DIR.iterdir()) if config.KB_DIR.exists() else []
    assert after == before


def test_summarize_keep_indexes(tmp_path: Path):
    path = _md_doc(tmp_path)
    result = summarize_path(path, keep=True, use_llm=False)
    assert result.kept is True
    assert result.keep_target is not None
    assert result.keep_target.exists()
    assert result.keep_target.is_symlink()


def test_ensure_citations_marks_missing():
    md = (
        "## 总结（最多三句话）\n一句话讲完。\n\n"
        "## 结构化要点\n"
        "- **好要点**：有依据 〔§安装 | p.2〕\n"
        "- **坏要点**：没有索引\n"
    )
    fixed, warnings = ensure_citations(md)
    assert "未定位到页/节" in fixed
    assert warnings
    assert citation_ok("- **好要点**：有依据 〔§安装 | p.2〕")
    assert not citation_ok("- **坏要点**：没有索引")


def _write_min_text_pdf(path: Path, lines: list[str]) -> None:
    """Write a tiny one-page PDF with Helvetica text (extractable by pypdf)."""
    # Simple PDF 1.4 with one content stream per logical line on one page.
    y = 750
    tj_ops = []
    for line in lines:
        safe = line.replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")
        tj_ops.append(f"BT /F1 12 Tf 50 {y} Td ({safe}) Tj ET")
        y -= 18
    stream = "\n".join(tj_ops)
    objects = [
        "1 0 obj<< /Type /Catalog /Pages 2 0 R >>endobj\n",
        "2 0 obj<< /Type /Pages /Kids [3 0 R] /Count 1 >>endobj\n",
        (
            "3 0 obj<< /Type /Page /Parent 2 0 R /MediaBox [0 0 612 792] "
            "/Contents 4 0 R /Resources<< /Font<< /F1 5 0 R >> >> >>endobj\n"
        ),
        f"4 0 obj<< /Length {len(stream)} >>stream\n{stream}\nendstream\nendobj\n",
        "5 0 obj<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>endobj\n",
    ]
    pdf = "%PDF-1.4\n"
    offsets = [0]
    for obj in objects:
        offsets.append(len(pdf))
        pdf += obj
    xref_pos = len(pdf)
    pdf += f"xref\n0 {len(objects) + 1}\n"
    pdf += "0000000000 65535 f \n"
    for off in offsets[1:]:
        pdf += f"{off:010d} 00000 n \n"
    pdf += (
        f"trailer<< /Size {len(objects) + 1} /Root 1 0 R >>\n"
        f"startxref\n{xref_pos}\n%%EOF\n"
    )
    path.write_bytes(pdf.encode("latin-1"))


def test_pdf_loader_page_markers(tmp_path: Path):
    out = tmp_path / "one.pdf"
    _write_min_text_pdf(out, ["Hello LocalAgent page one", "Second line on page"])
    doc = load_file(out)
    assert doc is not None
    assert "## [p.1]" in doc.text
    assert doc.metadata.get("page_count") == 1
    result = summarize_loaded(doc, use_llm=False)
    assert "p." in result.markdown or "§" in result.markdown
    assert result.page_count == 1


def test_document_too_long_without_allow_long(tmp_path: Path, monkeypatch: pytest.MonkeyPatch):
    monkeypatch.setattr(config, "SUMMARIZE_SHORT_MAX_CHARS", 100)
    path = tmp_path / "long.txt"
    path.write_text("字" * 200, encoding="utf-8")
    doc = load_file(path)
    assert doc is not None
    with pytest.raises(DocumentTooLongError):
        summarize_loaded(doc, use_llm=False, allow_long=False)


def test_summarize_path_long_doc_indexes_session(tmp_path: Path, monkeypatch: pytest.MonkeyPatch):
    monkeypatch.setattr(config, "SUMMARIZE_SHORT_MAX_CHARS", 100)
    monkeypatch.setattr(config, "SUMMARIZE_LLM_INPUT_CHARS", 80)
    path = tmp_path / "long.txt"
    # Enough text to exceed short max and trigger retrieval mode.
    path.write_text(("段落内容。" * 40 + "\n\n") * 5, encoding="utf-8")
    result = summarize_path(path, use_llm=False)
    assert result.session_source_key.startswith("sum:")
    assert result.char_count > config.SUMMARIZE_LLM_INPUT_CHARS
    assert result.uses_retrieval is True
    from localagent.summarize.document import format_document_context

    block = format_document_context(result)
    assert "速读卡" in block
    assert "检索" in block or "未塞入全文" in block


def test_format_document_context_includes_summary_and_source():
    from localagent.summarize.document import SummarizeResult, format_document_context

    result = SummarizeResult(
        markdown="## 总结（最多三句话）\n一句话。\n",
        path=Path("/tmp/x.md"),
        filename="x.md",
        char_count=10,
        annotated_text="## [§全文]\nhello world",
    )
    block = format_document_context(result)
    assert "当前文档" in block
    assert "一句话" in block
    assert "hello world" in block
    assert "未入库" in block
    assert "禁止建议用户再运行" in block


def test_session_index_and_source_filter(isolated_data, tmp_path: Path):
    from localagent.summarize.session_index import (
        format_retrieval_block,
        index_document_session,
        retrieve_document_chunks,
    )

    key = "sum:test:abc123"
    body = (
        "## [§架构]\n知识与推理分离通过独立模块实现，知识库用检索，推理用 Transformer。\n\n"
        "## [§评测]\n在科学任务上相对基线提升明显。\n"
    )
    n = index_document_session(key, body, title="demo")
    assert n >= 1
    hits = retrieve_document_chunks("知识与推理分离", source_key=key, top_k=3)
    assert hits
    assert all(
        (h.get("metadata") or {}).get("source_file") == key for h in hits
    )
    block = format_retrieval_block(hits, source_key=key)
    assert "当前文档检索结果" in block
    assert "知识" in block or "推理" in block


def test_unsupported_suffix(tmp_path: Path):
    path = tmp_path / "x.docx"
    path.write_text("nope", encoding="utf-8")
    with pytest.raises(Exception) as exc:
        summarize_path(path, use_llm=False)
    assert "不支持" in str(exc.value)


def test_markdown_suffix_accepted(tmp_path: Path):
    path = tmp_path / "notes.markdown"
    path.write_text(
        "# 章节\n\n"
        "这是一段用于速读测试的正文。\n",
        encoding="utf-8",
    )
    result = summarize_path(path, use_llm=False)
    assert "## 总结" in result.markdown
    assert result.filename == "notes.markdown"


def test_format_summarize_suffixes_includes_markdown():
    text = config.format_summarize_suffixes()
    assert ".markdown" in text
    assert ".md" in text
    assert text == config.format_summarize_suffixes(style="comma").replace(", ", " / ")


def test_should_enter_document_chat_respects_no_chat():
    from localagent.summarize.repl import should_enter_document_chat

    assert should_enter_document_chat(no_chat=True) is False


def test_not_kept_hint_reactive():
    assert not_kept_hint_if_asked("刚才一键总结的文档为啥没入库？") == KEEP_HINT
    assert not_kept_hint_if_asked("今天天气怎么样") is None


def test_xlsx_summarize(tmp_path: Path):
    from openpyxl import Workbook

    path = tmp_path / "t.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "预算"
    ws["A1"] = "项目"
    ws["B1"] = "金额"
    ws["A2"] = "服务器"
    ws["B2"] = 1000
    wb.save(path)
    result = summarize_path(path, use_llm=False)
    assert "## 总结" in result.markdown
    assert "预算" in result.markdown or "§" in result.markdown


def test_summarize_session_upsert_and_list(tmp_path: Path, monkeypatch: pytest.MonkeyPatch):
    monkeypatch.setattr(config, "DATA_DIR", tmp_path / "data")
    monkeypatch.setattr(config, "SUMMARIZE_SESSIONS_DIR", tmp_path / "data" / "summarize_sessions")
    monkeypatch.setattr(
        config,
        "SUMMARIZE_SESSIONS_INDEX",
        tmp_path / "data" / "summarize_sessions" / "index.json",
    )
    from localagent.summarize.sessions import (
        SummarizeSessionRecord,
        find_session_by_path,
        list_sessions,
        upsert_session,
    )

    path = _md_doc(tmp_path)
    record = SummarizeSessionRecord(
        id="sum-test01",
        path=str(path.resolve()),
        filename=path.name,
        mtime=path.stat().st_mtime,
        updated_at="",
        conversation_session_id="sum-test01",
        summary_md="## 总结\nhello\n",
        char_count=10,
        kept=False,
    )
    upsert_session(record)
    found = find_session_by_path(path)
    assert found is not None
    assert found.id == "sum-test01"
    assert list_sessions()[0].filename == path.name


def _summarize_args(tmp_path: Path, **overrides):
    import argparse

    if "paths" not in overrides and "path" not in overrides:
        overrides["paths"] = [str(_md_doc(tmp_path))]
    defaults = {
        "paths": [],
        "path": None,
        "no_chat": False,
        "keep": False,
        "heuristic": True,
        "provider": "auto",
        "model": None,
        "out": None,
        "list": False,
        "limit": 20,
        "resume": False,
        "force": False,
        "id": None,
        "no_ui": True,
        "no_prefetch": True,
        "refresh_segments": False,
        "retry_failed": False,
    }
    defaults.update(overrides)
    return argparse.Namespace(**defaults)


def test_cmd_summarize_auto_resumes_by_default(tmp_path: Path, monkeypatch: pytest.MonkeyPatch):
    from localagent.cli import cmd_summarize
    from localagent.summarize.sessions import SummarizeSessionRecord

    path = _md_doc(tmp_path)
    record = SummarizeSessionRecord(
        id="sum-auto01",
        path=str(path.resolve()),
        filename=path.name,
        mtime=path.stat().st_mtime,
        updated_at="",
        conversation_session_id="sum-auto01",
        summary_md="## 总结\nhello\n",
        char_count=10,
        kept=False,
    )
    resume_calls: list[tuple] = []

    monkeypatch.setattr(
        "localagent.summarize.sessions.find_session_by_path",
        lambda p: record if Path(p).resolve() == path.resolve() else None,
    )

    def _fake_resume(source, *, record, **kwargs):
        resume_calls.append((source, record.id, kwargs.get("refresh_cache")))
        return 0

    monkeypatch.setattr("localagent.cli._summarize_resume_one", _fake_resume)
    monkeypatch.setattr(
        "localagent.summarize.repl.should_enter_document_chat",
        lambda *, no_chat: not no_chat,
    )

    args = _summarize_args(tmp_path, paths=[str(path)], resume=False, force=False)
    assert cmd_summarize(args) == 0
    assert len(resume_calls) == 1
    assert resume_calls[0][1] == "sum-auto01"
    assert resume_calls[0][2] is False


def test_cmd_summarize_force_skips_resume_and_refreshes_cache(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
):
    from localagent.cli import cmd_summarize
    from localagent.summarize.sessions import SummarizeSessionRecord

    path = _md_doc(tmp_path)
    record = SummarizeSessionRecord(
        id="sum-force01",
        path=str(path.resolve()),
        filename=path.name,
        mtime=path.stat().st_mtime,
        updated_at="",
        conversation_session_id="sum-force01",
        summary_md="## 总结\nhello\n",
        char_count=10,
        kept=False,
    )
    resume_calls: list[tuple] = []
    summarize_calls: list[tuple] = []

    monkeypatch.setattr(
        "localagent.summarize.sessions.find_session_by_path",
        lambda p: record if Path(p).resolve() == path.resolve() else None,
    )
    monkeypatch.setattr(
        "localagent.cli._summarize_resume_one",
        lambda *a, **k: resume_calls.append((a, k)) or 0,
    )

    def _fake_summarize(
        source,
        *,
        keep,
        use_llm,
        refresh_cache,
        retry_failed,
        translate=None,
        model_choice=None,
    ):
        from localagent.summarize.document import SummarizeResult

        summarize_calls.append((source, refresh_cache, retry_failed))
        return SummarizeResult(
            markdown="## 总结\nnew\n",
            path=Path(source),
            filename=Path(source).name,
            char_count=10,
            used_llm=False,
        )

    monkeypatch.setattr("localagent.summarize.document.summarize_path", _fake_summarize)
    monkeypatch.setattr("localagent.summarize.repl.enter_summarize_interactive", lambda *a, **k: 0)
    monkeypatch.setattr(
        "localagent.summarize.repl.should_enter_document_chat",
        lambda *, no_chat: not no_chat,
    )

    args = _summarize_args(tmp_path, paths=[str(path)], force=True)
    assert cmd_summarize(args) == 0
    assert resume_calls == []
    assert len(summarize_calls) == 1
    assert summarize_calls[0][1] is True


def test_cmd_summarize_no_chat_does_not_auto_resume(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
):
    from localagent.cli import cmd_summarize
    from localagent.summarize.sessions import SummarizeSessionRecord

    path = _md_doc(tmp_path)
    record = SummarizeSessionRecord(
        id="sum-nochat01",
        path=str(path.resolve()),
        filename=path.name,
        mtime=path.stat().st_mtime,
        updated_at="",
        conversation_session_id="sum-nochat01",
        summary_md="## 总结\nhello\n",
        char_count=10,
        kept=False,
    )

    monkeypatch.setattr(
        "localagent.summarize.sessions.find_session_by_path",
        lambda p: record if Path(p).resolve() == path.resolve() else None,
    )
    monkeypatch.setattr(
        "localagent.cli._summarize_resume_one",
        lambda *a, **k: (_ for _ in ()).throw(AssertionError("should not resume")),
    )

    args = _summarize_args(tmp_path, paths=[str(path)], no_chat=True)
    assert cmd_summarize(args) == 0


def test_cmd_summarize_multi_requires_no_chat(tmp_path: Path):
    import argparse

    from localagent.cli import cmd_summarize

    a = _md_doc(tmp_path)
    b = tmp_path / "b.md"
    b.write_text("# B\n\nok\n", encoding="utf-8")
    args = argparse.Namespace(
        paths=[str(a), str(b)],
        path=None,
        no_chat=False,
        keep=False,
        heuristic=True,
        provider="auto",
        out=None,
        list=False,
        limit=20,
        resume=False,
        force=False,
        id=None,
        no_ui=True,
        no_prefetch=True,
        refresh_segments=False,
        retry_failed=False,
    )
    assert cmd_summarize(args) == 1


def test_cmd_summarize_no_chat_batch(tmp_path: Path):
    import argparse

    from localagent.cli import cmd_summarize

    a = _md_doc(tmp_path)
    b = tmp_path / "b.md"
    b.write_text("# 第二份\n\n内容很短。\n", encoding="utf-8")
    out = tmp_path / "out.md"
    args = argparse.Namespace(
        paths=[str(a), str(b)],
        path=None,
        no_chat=True,
        keep=False,
        heuristic=True,
        provider="auto",
        out=str(out),
        list=False,
        limit=20,
        resume=False,
        force=False,
        id=None,
        no_ui=True,
        no_prefetch=True,
        refresh_segments=False,
        retry_failed=False,
    )
    assert cmd_summarize(args) == 0
    text = out.read_text(encoding="utf-8")
    assert "brief.md" in text or "产品概述" in text
    assert "第二份" in text or "b.md" in text


def test_heuristic_summarize_appends_source_footer(tmp_path: Path):
    path = _md_doc(tmp_path)
    result = summarize_path(path, keep=False, use_llm=False)
    assert "*摘要 via 本地启发式*" in result.markdown
    assert result.source is not None
    assert result.source.via == "heuristic"


def test_llm_summarize_passes_model_choice(tmp_path: Path, monkeypatch: pytest.MonkeyPatch):
    from localagent.models.router import ChatResult
    from localagent.summarize.model_choice import SummarizeModelChoice

    path = _md_doc(tmp_path)
    calls: list[dict] = []

    class FakeRouter:
        def chat_with_meta(self, messages, **kwargs):
            calls.append(kwargs)
            return ChatResult(
                text=(
                    "## 总结（最多三句话）\n"
                    "一句话。\n\n"
                    "## 结构化要点\n"
                    "- **点**：内容 — 依据：原文 〔§安装〕\n"
                ),
                provider="ollama",
                model="qwen3:8b",
            )

    monkeypatch.setattr(
        "localagent.models.router.get_model_router",
        lambda: FakeRouter(),
    )
    choice = SummarizeModelChoice(provider="ollama", model="qwen3:8b")
    result = summarize_path(path, keep=False, use_llm=True, model_choice=choice)
    assert calls
    assert calls[0]["prefer"] == "ollama"
    assert calls[0]["model"] == "qwen3:8b"
    assert "*摘要 via ollama/qwen3:8b*" in result.markdown
    assert result.source is not None
    assert result.source.via == "llm"


_VALID_LLM_CARD = (
    "## 总结（最多三句话）\n"
    "一句话。\n\n"
    "## 结构化要点\n"
    "- **点**：内容 — 依据：原文 〔§安装〕\n"
)


def test_llm_summarize_retries_ollama_after_primary_failure(
    monkeypatch: pytest.MonkeyPatch,
):
    from localagent.models.router import ChatResult
    from localagent.summarize.document import _llm_summarize
    from localagent.summarize.model_choice import SummarizeModelChoice

    calls: list[str | None] = []

    class FakeRouter:
        def is_ollama_available(self) -> bool:
            return True

        def chat_with_meta(self, messages, **kwargs):
            calls.append(kwargs.get("prefer"))
            if kwargs.get("prefer") == "openrouter":
                raise RuntimeError("cloud down")
            return ChatResult(
                text=_VALID_LLM_CARD,
                provider="ollama",
                model="qwen3:8b",
            )

    monkeypatch.setattr(
        "localagent.models.router.get_model_router",
        lambda: FakeRouter(),
    )
    text, source = _llm_summarize(
        "## [§安装]\n内容。",
        filename="t.md",
        model_choice=SummarizeModelChoice(provider="openrouter"),
    )
    assert text
    assert source is not None
    assert source.provider == "ollama"
    assert calls == ["openrouter", "ollama"]


def test_llm_summarize_no_heuristic_after_ollama_retries_exhausted(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    from localagent.summarize.document import SUMMARIZE_OLLAMA_RETRIES

    calls: list[int] = []

    class FakeRouter:
        def is_ollama_available(self) -> bool:
            return True

        def chat_with_meta(self, messages, **kwargs):
            calls.append(1)
            raise RuntimeError("all down")

    monkeypatch.setattr(
        "localagent.models.router.get_model_router",
        lambda: FakeRouter(),
    )
    path = _md_doc(tmp_path)
    result = summarize_path(path, keep=False, use_llm=True)
    assert "本地启发式" not in result.markdown
    assert result.markdown == ""
    assert "模型摘要失败" in result.warnings
    assert len(calls) == 1 + SUMMARIZE_OLLAMA_RETRIES
